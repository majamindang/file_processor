import re, openpyxl
from openpyxl.styles import Alignment, Font
from datetime import datetime

def dbp_statement_txt_excel(input_file, output_file):
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FILE"

        rows = open(input_file, 'r').read().splitlines()

        EXCEL_NUMBER_FORMAT = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'

        def convert_to_date(date):
            months = ['', 'jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
            day, month, year = date.strip().split(" ")
            this_date = datetime(int(year), months.index(month.lower()), int(day))
            return this_date.date()

        def convert_to_number(amount):
            if not amount.strip(): return 0
            return float(amount.strip().replace(',', ''))

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 5
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 22
        ws.column_dimensions["E"].width = 22
        ws.column_dimensions["F"].width = 22
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 6

        row_count = 2

        for _r, i in enumerate(['DATE', 'TC', 'REF/CHECK NO', 'DEBIT', 'CREDIT', 'BALANCE', 'REF_DATE', 'TX BR']):
            ws.cell(1, _r + 1).value = i
            ws.cell(1, _r + 1).font = Font(bold = True)

            if(_r in [3, 4, 5]):
                ws.cell(1, _r + 1).alignment = Alignment(horizontal='right')

        total_debit = 0
        total_credit = 0

        for row in rows:
            # pattern = r"^\s+(\d{2} \w{3} \d{4})\s+(\w{3})\s+([a-zA-Z0-9#]*)?\s+([0-9\.\,]+)?\s+([0-9\.\,]+)?\s+([0-9\.\,]+)?\s+(\d{2} \w{3} \d{4})\s+(\d{4})\n?$"
            pattern = r"^\s{2}(.{11})\s{3}(.{3})\s{3}(.{16})\s{3}(.{20})\s{3}(.{20})\s{3}(.{20})\s{3}(.{11})\s{3}(.{4})\n?$"

            if re.match(pattern, row):
                this_line = re.findall(pattern, row)[0]
                date, tc, ref, debit, credit, balance, ref_date, tx_br = this_line
                ws.cell(row_count, 1).value = convert_to_date(date)
                ws.cell(row_count, 2).value = tc.strip()
                ws.cell(row_count, 3).value = ref.strip()
                ws.cell(row_count, 4).value = convert_to_number(debit)
                ws.cell(row_count, 4).number_format = EXCEL_NUMBER_FORMAT
                ws.cell(row_count, 5).value = convert_to_number(credit)
                ws.cell(row_count, 5).number_format = EXCEL_NUMBER_FORMAT
                ws.cell(row_count, 6).value = convert_to_number(balance)
                ws.cell(row_count, 6).number_format = EXCEL_NUMBER_FORMAT
                ws.cell(row_count, 7).value = convert_to_date(ref_date)
                ws.cell(row_count, 8).value = tx_br.strip()

                total_debit += ws.cell(row_count, 4).value
                total_credit += ws.cell(row_count, 5).value

                row_count += 1

        ws.cell(row_count + 1, 1).value = "TOTAL"
        ws.cell(row_count + 1, 1).font = Font(bold = True)
        ws.cell(row_count + 1, 1).alignment = Alignment(horizontal='right')
        ws.cell(row_count + 1, 4).value = total_debit
        ws.cell(row_count + 1, 4).number_format = EXCEL_NUMBER_FORMAT
        ws.cell(row_count + 1, 4).font = Font(bold = True)
        ws.cell(row_count + 1, 5).value = total_credit
        ws.cell(row_count + 1, 5).number_format = EXCEL_NUMBER_FORMAT
        ws.cell(row_count + 1, 5).font = Font(bold = True)

        wb.save(output_file)

    except Exception as e:
        return (False, e.strerror)

    return (True, f"Successfully processed the file and saved as '{output_file}' ")